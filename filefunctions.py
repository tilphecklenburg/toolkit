#-------import required modules for functions below----------------------
from openpyxl import Workbook, load_workbook
import os
import datetime
#------------------------------------------------------------------------

#-------create a new Excel workbook using OpenPyxl-----------------------


def createexceldoc(filename):
	filename = Workbook()
	print("workbook created")
	return filename
#------------------------------------------------------------------------	

#-------create a new Excel worksheet using OpenPyxl----------------------


def createexcelsheet(workbookname, sheetname):
	excelsheet = workbookname.create_sheet(title=sheetname)
	print("sheet created")
	return excelsheet
#------------------------------------------------------------------------

#-------search for string in a .txt file---------------------------------


def searchtextfile(filename, searchphrase, reportsheet):
	rowcount = 1
	file = open(filename, "r+")
	for line in file.readlines():
		if searchphrase in line:
			reportsheet['A' + str(rowcount)] = searchphrase
			reportsheet['B' + str(rowcount)] = line
			rowcount += 1
#------------------------------------------------------------------------	

#-------open text file and turn into a list------------------------------


def parsetextfiletolist(filename):
	file = open(filename,"r+")
	list = file.read()
	list = list.split()
	return list
#------------------------------------------------------------------------

#-------gather directory listing and parse/search each file within dir---


def searchdirforstring(dir, string, reportsheet):
	rowcount = 1
	print("Note: This function can only search text files at the moment")
	dirlist = os.listdir(dir)
	reportsheet['A' + str(rowcount)] = "Search string:"
	reportsheet['B' + str(rowcount)] = "Found in file:"
	reportsheet['C' + str(rowcount)] = "Config line:"
	rowcount = 2
	for file in dirlist:
		if ".txt" in str(file):
			searchdata = open(str(dir + file),"r+")
			searchdata = str(searchdata.read())
			searchdata = searchdata.split("\n")
			for line in searchdata:
				if string.lower() in line.lower():
					reportsheet['A' + str(rowcount)] = string
					reportsheet['B' + str(rowcount)] = file
					reportsheet['C' + str(rowcount)] = line
					rowcount += 1
	matches = rowcount - 2
	print("Search complete - %s matches found - see" % str(matches) + str(string) + "results.xlsx in toolkit folder") 
	return reportsheet, rowcount
#------------------------------------------------------------------------		

#-------gather directory listing and parse/search each file within dir for multiple strings provided---


def searchdirformultistring(dir, hosts, reportsheet):
	hosts = hosts.read()
	hosts = hosts.split("\n")
	rowcount = 1
	print("Note: This function can only search text files at the moment")
	dirlist = os.listdir(dir)
	reportsheet['A' + str(rowcount)] = "Search string:"
	reportsheet['B' + str(rowcount)] = "Found on device:"
	reportsheet['C' + str(rowcount)] = "Config line:"
	rowcount = 2
	for file in dirlist:
		if ".txt" in str(file):
			searchdata = open(str(dir + file),"r+")
			searchdata = str(searchdata.read())
			searchdata = searchdata.split("\n")
			for line in searchdata:
				for string in hosts:
					if string.lower() in line.lower():
						reportsheet['A' + str(rowcount)] = string.upper()
						reportsheet['B' + str(rowcount)] = file
						reportsheet['C' + str(rowcount)] = line
						rowcount += 1
	matches = rowcount - 2
	print("Search complete - %s matches found - see results.xlsx in Config Searcher folder" % matches) 
	return reportsheet, rowcount
#------------------------------------------------------------------------


#-------take exceldoc, excelsheet, ipcolumn and provide ping output in resultscolumn--------------
def pinghostsinexceldoc(exceldoc, excelsheet, ipcolumn, resultscolumn):
	#make sure we have all the required variables, gather them if not
	if exceldoc=='':
		exceldoc=input("Please provide the name of the Excel doc: ")
	if excelsheet=='':
		excelsheet=input("Please provide the name of the Excel sheet with IPs: ")
	if ipcolumn=='':
		ipcolumn=input("Please provide the column with IP addresses in it: ")
	if resultscolumn=='':
		resultscolumn=input("Please provide the column that results should be stored in: ")
	try:
		workbook = load_workbook(exceldoc)
	except:
		print("Unable to load workbook, make sure script is running in same directory as XLSX file")
	try:
		worksheet = workbook.get_sheet_by_name(excelsheet)
	except:
		print("Unable to load worksheet")
	rowcount = 1
	print('trying some pings 1')
	try:
		for cell in worksheet[ipcolumn]:
			cell_length = len(str(cell.value))
			print(cell_length)
			if rowcount > 1 and int(cell_length) < 16:
				print('attempting pings to %s' % str(cell.value))
				result = os.system('ping -w 1000 -n 1 ' + str(cell.value))
				print(result)
				if result == 0:
					print('%s is reachable' % str(cell.value))
					worksheet[str(resultscolumn) + str(rowcount)] = 'Y - last response %s' % str(datetime.datetime.now())
				else:
					print('%s is not reachable' % str(cell.value))
					worksheet[str(resultscolumn) + str(rowcount)] = 'N - last attempt %s' % str(datetime.datetime.now())
				rowcount += 1
			else:
				print('something wrong with cell, skipping. ensure that a single IP has been specified per cell in selected column')
				rowcount += 1
	except:
		workbook.save('AltaPACS - Master Device List.xlsx')
		workbook.close()
#------------------------------------------------------------------------
	workbook.save('AltaPACS - Master Device List.xlsx')
	workbook.close()

